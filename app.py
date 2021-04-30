from flask import Flask, render_template, request
from app.config.middleware import checkLogin
from app.services import user, karakteristik, spesifik, misc, jenis_pesawat, dataset, svmv
import os
from werkzeug.utils import secure_filename
import pandas as pd
import numpy as np
from sklearn import svm
import openpyxl
from app.config import db
import flask_excel as excel

app = Flask(__name__)

ALLOWED_EXTENSION = set(['xlsx'])
app.config['UPLOAD_FOLDER'] = 'uploads'

excel.init_excel(app)


@app.route("/")
@checkLogin
def index():
    return misc.index()


@app.route("/login")
def login():
    return misc.login()


@app.route("/doLogin", methods=['POST'])
def doLogin():
    return misc.doLogin(request.form)


@app.route("/logout")
def logout():
    return misc.logout()

# ---------------- User Route --------------- #


@app.route('/users')
@checkLogin
def user_index():
    return user.index()


@app.route('/user/store', methods=['POST'])
@checkLogin
def user_store():
    return user.store(request.form)


@app.route("/user/<int:id>/update", methods=['POST'])
@checkLogin
def user_update(id):
    return user.update(request.form, id)


@app.route("/user/<int:id>/delete", methods=['POST'])
@checkLogin
def user_delete(id):
    return user.delete(id)

# ---------------- END User Route --------------- #


@app.route("/klasifikasisvm")
@checkLogin
def svm_index():
    return svmv.index()


@app.route("/cekklasifikasisvm", methods=["POST"])
@checkLogin
def cek_klasifikasi_svm():
    return svmv.check(request.form)

# ---------------- Ciri-ciri Route --------------- #


@app.route('/karakteristik')
@checkLogin
def karakteristik_index():
    return karakteristik.index()


@app.route('/karakteristik/store', methods=['POST'])
@checkLogin
def karakteristik_store():
    return karakteristik.store(request.form)


@app.route("/karakteristik/<int:id>/update", methods=['POST'])
@checkLogin
def karakteristik_update(id):
    return karakteristik.update(request.form, id)


@app.route("/karakteristik/<int:id>/delete", methods=['POST'])
@checkLogin
def karakteristik_delete(id):
    return karakteristik.delete(id)

# ---------------- END Ciri-ciri Route --------------- #

# ---------------- Spesifik Route --------------- #


@app.route('/spesifik')
@checkLogin
def spesifik_index():
    return spesifik.index()


@app.route('/spesifik/store', methods=['POST'])
@checkLogin
def spesifik_store():
    return spesifik.store(request.form)


@app.route("/spesifik/<int:id>/update", methods=['POST'])
@checkLogin
def spesifik_update(id):
    return spesifik.update(request.form, id)


@app.route("/spesifik/<int:id>/delete", methods=['POST'])
@checkLogin
def spesifik_delete(id):
    return spesifik.delete(id)

# ---------------- END Spesifik Route --------------- #

# ---------------- Jenis Pesawat Route --------------- #


@app.route('/jenis_pesawat')
@checkLogin
def jenis_pesawat_index():
    return jenis_pesawat.index()


@app.route('/jenis_pesawat/store', methods=['POST'])
@checkLogin
def jenis_pesawat_store():
    return jenis_pesawat.store(request.form)


@app.route("/jenis_pesawat/<int:id>/update", methods=['POST'])
@checkLogin
def jenis_pesawat_update(id):
    return jenis_pesawat.update(request.form, id)


@app.route("/jenis_pesawat/<int:id>/delete", methods=['POST'])
@checkLogin
def jenis_pesawat_delete(id):
    return jenis_pesawat.delete(id)

# ---------------- END Jenis Pesawat Route --------------- #

# ---------------- Import Route --------------- #


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSION


uploads_dir = os.path.join(app.instance_path, 'uploads')


@app.route('/vupload')
@checkLogin
def vupload():
    return render_template('pages/import/index.html')


@app.route('/view', methods=['GET', 'POST'])
@checkLogin
def view():
    if request.method == 'POST':

        file = request.files['file']

        if 'file' not in request.files:
            return render_template('pages/import/index.html')

        if file.filename == '':
            return render_template('pages/import/index.html')

        if file and allowed_file(file.filename):
            column_list = []
            df_column = pd.read_excel(file).columns
            for i in df_column:
                column_list.append(i)

            converter = {col: str for col in column_list}
            data_xls = pd.read_excel(file, converters=converter)
            # print(data_xls.head())
            return dataset.upload(data_xls)

            # filename = secure_filename(file.filename)
            # file.save(os.path.join(uploads_dir,  filename))
            # loc = uploads_dir + '/' + filename
            # book = openpyxl.load_workbook(loc)
            # sheet = book.sheet_by_index(0)
            # conn = db.conn()
            # with conn.cursor() as cursor:
            #     sql = "INSERT INTO tbl_pesawat (`nama_pesawat`,`id_jenis_pesawat`,`id_jenis_sayap`,`id_jenis_penempatan_sayap`,`id_jumlah_sayap`,`id_arah_sayap`,`id_jenis_mesin`,`id_jumlah_mesin`,`id_posisi_mesin`,`id_badan_pesawat`,`id_jenis_ekor`,`id_jenis_landing_gear`,`id_canard`,`id_persenjataan`) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,)"

            #     for r in range(1, sheet.nrows):
            #         nama_pesawat = sheet.cell(r, 0).value
            #         id_jenis_pesawat = sheet.cell(r, 1).value
            #         id_jenis_sayap = sheet.cell(r, 2).value
            #         id_jenis_penempatan_sayap = sheet.cell(r, 3).value
            #         id_jumlah_sayap = sheet.cell(r, 4).value
            #         id_arah_sayap = sheet.cell(r, 5).value
            #         id_jenis_mesin = sheet.cell(r, 6).value
            #         id_jumlah_mesin = sheet.cell(r, 7).value
            #         id_posisi_mesin = sheet.cell(r, 8).value
            #         id_badan_pesawat = sheet.cell(r, 9).value
            #         id_jenis_ekor = sheet.cell(r, 10).value
            #         id_jenis_landing_gear = sheet.cell(r, 11).value
            #         id_canard = sheet.cell(r, 12).value
            #         id_persenjataan = sheet.cell(r, 13).value

            #     cursor.execute(sql, (nama_pesawat, id_jenis_pesawat, id_jenis_sayap, id_jenis_penempatan_sayap, id_jumlah_sayap, id_arah_sayap,
            #                          id_jenis_mesin, id_jumlah_mesin, id_posisi_mesin, id_badan_pesawat, id_jenis_ekor, id_jenis_landing_gear, id_canard, id_persenjataan))

            # conn.commit()
            # conn.close()
            # return render_template("pages/import/view.html")
# ---------------- END Import Route --------------- #


@app.route('/dataset')
@checkLogin
def dataset_index():
    return dataset.index()


@app.route('/dataset/create')
@checkLogin
def dataset_create():
    return dataset.create()


@app.route('/dataset/store', methods=['POST'])
@checkLogin
def dataset_store():
    # print(request.form)
    # return request.form
    return dataset.store(request.form)


app.secret_key = '3RDLwwtFttGSxkaDHyFTmvGytBJ2MxWT8ynWm2y79G8jm9ugYxFFDPdHcBBnHp6E'
app.config['SESSION_TYPE'] = 'filesystem'


@app.context_processor
def inject_stage_and_region():
    return dict(APP_NAME=os.environ.get("APP_NAME"),
                APP_AUTHOR=os.environ.get("APP_AUTHOR"),
                APP_TITLE=os.environ.get("APP_TITLE"))


if __name__ == "__main__":
    app.run(debug=True)
